from Sumin.spread import *
import openpyxl

class Write:
    def __init__(self,old_spread,new_spread,file):
        self.old_obj = old_spread.s_obj
        self.old_match = old_spread.s_match
        self.old_issue = old_spread.s_issue

        self.new_obj = new_spread.s_obj
        self.new_match = new_spread.s_match
        self.new_issue = new_spread.s_issue
        self.file = file
        self.excel_write()
        pass

    def excel_write(self):
        wb = openpyxl.Workbook()
        ws1 = wb.create_sheet("old_obj")
        ws2 = wb.create_sheet("old_issue")
        ws3 = wb.create_sheet("old_match")
        ws4 = wb.create_sheet("new_obj")
        ws5 = wb.create_sheet("new_issue")
        ws6 = wb.create_sheet("new_match")

        ws1.append([None])
        for t in range(len(self.old_obj)):
            ws1.append(self.old_obj[t])

        ws2.append([None])
        for t in range(len(self.old_issue)):
            ws2.append(self.old_issue[t])

        ws3.append([None])
        for t in range(len(self.old_match)):
            ws3.append(self.old_match[t])

        ws4.append([None])
        for t in range(len(self.new_obj)):
            ws4.append(self.new_obj[t])

        ws5.append([None])
        for t in range(len(self.new_issue)):
            ws5.append(self.new_issue[t])

        ws6.append([None])
        for t in range(len(self.new_match)):
            ws6.append(self.new_match[t])
        wb.save(".\\Output\\" + self.file + ".xlsx")
        pass

    pass